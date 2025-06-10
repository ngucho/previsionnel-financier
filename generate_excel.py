from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

wb = Workbook()
ws = wb.active
ws.title = "Prévisionnel"

headers = ["Catégorie", "Montant (€)"]
ws.append(headers)
ws["A1"].font = Font(bold=True)
ws["B1"].font = Font(bold=True)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 20

sample_data = [
    ("Recettes", 1000000),
    ("Coûts", 800000),
    ("Bénéfice", "=B2-B3"),
]

for row in sample_data:
    ws.append(row)

for cell in ws[1]:
    cell.alignment = Alignment(horizontal="center")

wb.save("financial_template.xlsx")
print("fichier 'financial_template.xlsx' créé")
